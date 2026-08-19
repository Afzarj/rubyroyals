import os
import io
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

app = Flask(__name__)
app.secret_key = (
    os.environ.get("FLASK_SECRET")
    or os.environ.get("SECRET_KEY")
    or "dev_secret"
)

# --- Database Config ---
# Use SQLite locally, PostgreSQL on Render
database_url = os.environ.get("DATABASE_URL", "sqlite:///pledges.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
if database_url.startswith("postgresql://"):
    database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# --- Models ---
class Pledge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bill_no = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    gender = db.Column(db.String(20))
    age = db.Column(db.Integer)
    father_name = db.Column(db.String(100))
    family_name = db.Column(db.String(100))
    kids_names = db.Column(db.String(200))
    education = db.Column(db.String(100))
    occupation = db.Column(db.String(100))
    address = db.Column(db.String(200))
    phone = db.Column(db.String(10), nullable=False)
    alt_number = db.Column(db.String(10))
    aadhar = db.Column(db.String(12), nullable=False)
    hc_claim_form = db.Column(db.String(50))
    intro = db.Column(db.String(20))
    num_ornaments = db.Column(db.Integer)
    ornaments_details = db.Column(db.Text)  # JSON string of ornaments
    pledge_date = db.Column(db.Date)
    total_amount = db.Column(db.Float)
    total_grams = db.Column(db.Float)
    return_jewellery = db.Column(db.Float, default=0)
    balance_jewellery = db.Column(db.Float)
    repayment = db.Column(db.String(20))
    repayment_details = db.Column(db.Text)  # JSON string of repayments
    repayment_total_amount = db.Column(db.Float, default=0)
    remarks = db.Column(db.Text)

# --- Auth Decorator ---
def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get("is_admin"):
            return fn(*args, **kwargs)
        flash("Admin login required", "danger")
        return redirect(url_for("admin_login", next=request.path))
    return wrapper

# --- Routes ---
@app.route("/", methods=["GET", "POST"])
def pledge_form():
    if request.method == "POST":
        bill_no = request.form.get("BillNo")
        phone = request.form.get("PhoneNumber")
        aadhar = request.form.get("AadharNumber")

        # Validations
        if Pledge.query.filter_by(bill_no=bill_no).first():
            return render_template("form.html", error=f"Duplicate Bill No {bill_no} already exists")

        if not phone.isdigit() or len(phone) != 10:
            return render_template("form.html", error="Phone number must be exactly 10 digits")

        if not aadhar.isdigit() or len(aadhar) != 12:
            return render_template("form.html", error="Aadhar number must be exactly 12 digits")
        repayment_total_amount = sum(
            float(value or 0)
            for key, value in request.form.items()
            if key.startswith("RepayAmount")
        )
        pledge_date_str = request.form.get("PledgeDate")
        pledge_date = None
        if pledge_date_str:
            try:
                pledge_date = datetime.strptime(pledge_date_str, "%Y-%m-%d").date()
            except ValueError:
                return render_template("form.html", error="Invalid date format")
        pledge = Pledge(
            bill_no=bill_no,
            name=request.form.get("Name"),
            gender=request.form.get("Gender"),
            age=request.form.get("Age"),
            father_name=request.form.get("FatherName"),
            family_name=request.form.get("FamilyName"),
            kids_names=",".join([request.form.get("KidsNames1",""), request.form.get("KidsNames2",""), request.form.get("KidsNames3","")]),
            education=request.form.get("Education"),
            occupation=request.form.get("Occupation"),
            address=request.form.get("Address"),
            phone=phone,
            alt_number=request.form.get("AltNumber"),
            aadhar=aadhar,
            hc_claim_form=request.form.get("HCClaimFormNumber"),
            intro=request.form.get("Intro"),
            num_ornaments=request.form.get("NumOrnaments"),
            ornaments_details=str({k:v for k,v in request.form.items() if "Ornament" in k or "Grams" in k}),
            pledge_date=pledge_date,
            total_amount=float(request.form.get("TotalAmounts")),
            total_grams=float(request.form.get("TotalGrams")),
            return_jewellery=float(request.form.get("ReturnJewellery")),
            balance_jewellery=float(request.form.get("BalanceJewellery")),
            repayment=request.form.get("Repayment"),
            repayment_details=str({k:v for k,v in request.form.items() if "Repay" in k}),
            repayment_total_amount=round(repayment_total_amount, 2),
            remarks=request.form.get("Remarks")
        )
        db.session.add(pledge)
        db.session.commit()
        return redirect(url_for("success"))

    return render_template("form.html")

@app.route("/success")
def success():
    return render_template("success.html")

@app.route("/export")
@admin_required
def export_excel():
    pledges = Pledge.query.all()

    # Build a list of dicts with ALL fields
    data = []
    for p in pledges:
        data.append({
            "ID": p.id,
            "BillNo": p.bill_no,
            "Name": p.name,
            "Gender": p.gender,
            "Age": p.age,
            "FatherName": p.father_name,
            "FamilyName": p.family_name,
            "KidsNames": p.kids_names,
            "Education": p.education,
            "Occupation": p.occupation,
            "Address": p.address,
            "Phone": p.phone,
            "AltNumber": p.alt_number,
            "Aadhar": p.aadhar,
            "HCClaimForm": p.hc_claim_form,
            "Intro": p.intro,
            "NumOrnaments": p.num_ornaments,
            "OrnamentsDetails": p.ornaments_details,
            "PledgeDate": p.pledge_date,
            "TotalAmount": p.total_amount,
            "TotalGrams": p.total_grams,
            "ReturnJewellery": p.return_jewellery,
            "BalanceJewellery": p.balance_jewellery,
            "Repayment": p.repayment,
            "RepaymentDetails": p.repayment_details,
            "RepaymentTotalAmount": p.repayment_total_amount or 0,
            "Remarks": p.remarks
        })

    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pledges"

    if data:
        headers = list(data[0].keys())
        worksheet.append(headers)
        for row in data:
            worksheet.append([row[header] for header in headers])

    workbook.save(output)
    output.seek(0)
    filename = f"pledges_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True)


def _excel_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _excel_float(value, field_name, row_number, default=0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Row {row_number}: {field_name} must be a number") from error


def _excel_date(value, field_name, row_number):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date() if hasattr(value, "hour") else value
    for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), date_format).date()
        except ValueError:
            continue
    raise ValueError(f"Row {row_number}: {field_name} must be a valid date")


@app.route("/admin/import", methods=["POST"])
@admin_required
def import_excel():
    uploaded_file = request.files.get("excel_file")
    if not uploaded_file or not uploaded_file.filename:
        flash("Please choose an Excel file to import.", "danger")
        return redirect(url_for("admin_dashboard"))
    if not uploaded_file.filename.lower().endswith(".xlsx"):
        flash("Only .xlsx Excel files are supported.", "danger")
        return redirect(url_for("admin_dashboard"))

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [_excel_text(value) for value in next(rows, ())]
        required_headers = {"BillNo", "Name", "Phone", "Aadhar"}
        missing_headers = sorted(required_headers - set(headers))
        if missing_headers:
            raise ValueError(f"Missing required columns: {', '.join(missing_headers)}")

        header_indexes = {header: index for index, header in enumerate(headers) if header}
        existing_bill_numbers = {pledge.bill_no for pledge in Pledge.query.all()}
        imported_bill_numbers = set()
        imported_pledges = []

        for row_number, values in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = {
                header: values[index] if index < len(values) else None
                for header, index in header_indexes.items()
            }
            bill_no = _excel_text(row.get("BillNo"))
            name = _excel_text(row.get("Name"))
            phone = _excel_text(row.get("Phone"))
            aadhar = _excel_text(row.get("Aadhar"))
            if not bill_no or not name or not phone or not aadhar:
                raise ValueError(f"Row {row_number}: BillNo, Name, Phone, and Aadhar are required")
            if bill_no in existing_bill_numbers or bill_no in imported_bill_numbers:
                raise ValueError(f"Row {row_number}: BillNo {bill_no} already exists")
            if not phone.isdigit() or len(phone) != 10:
                raise ValueError(f"Row {row_number}: Phone must be exactly 10 digits")
            if not aadhar.isdigit() or len(aadhar) != 12:
                raise ValueError(f"Row {row_number}: Aadhar must be exactly 12 digits")

            imported_pledges.append(Pledge(
                bill_no=bill_no,
                name=name,
                gender=_excel_text(row.get("Gender")) or None,
                age=int(_excel_float(row.get("Age"), "Age", row_number)) if row.get("Age") not in (None, "") else None,
                father_name=_excel_text(row.get("FatherName")) or None,
                family_name=_excel_text(row.get("FamilyName")) or None,
                kids_names=_excel_text(row.get("KidsNames")) or None,
                education=_excel_text(row.get("Education")) or None,
                occupation=_excel_text(row.get("Occupation")) or None,
                address=_excel_text(row.get("Address")) or None,
                phone=phone,
                alt_number=_excel_text(row.get("AltNumber")) or None,
                aadhar=aadhar,
                hc_claim_form=_excel_text(row.get("HCClaimForm")) or None,
                intro=_excel_text(row.get("Intro")) or None,
                num_ornaments=int(_excel_float(row.get("NumOrnaments"), "NumOrnaments", row_number)) if row.get("NumOrnaments") not in (None, "") else None,
                ornaments_details=_excel_text(row.get("OrnamentsDetails")) or None,
                pledge_date=_excel_date(row.get("PledgeDate"), "PledgeDate", row_number),
                total_amount=_excel_float(row.get("TotalAmount"), "TotalAmount", row_number),
                total_grams=_excel_float(row.get("TotalGrams"), "TotalGrams", row_number),
                return_jewellery=_excel_float(row.get("ReturnJewellery"), "ReturnJewellery", row_number),
                balance_jewellery=_excel_float(row.get("BalanceJewellery"), "BalanceJewellery", row_number),
                repayment=_excel_text(row.get("Repayment")) or None,
                repayment_details=_excel_text(row.get("RepaymentDetails")) or None,
                repayment_total_amount=_excel_float(row.get("RepaymentTotalAmount"), "RepaymentTotalAmount", row_number),
                remarks=_excel_text(row.get("Remarks")) or None
            ))
            imported_bill_numbers.add(bill_no)

        if not imported_pledges:
            raise ValueError("The workbook does not contain any data rows")
        db.session.add_all(imported_pledges)
        db.session.commit()
        flash(f"Imported {len(imported_pledges)} pledge record(s) successfully.", "success")
    except (InvalidFileException, ValueError) as error:
        db.session.rollback()
        flash(f"Import failed: {error}", "danger")
    except Exception:
        db.session.rollback()
        flash("Import failed. Check that the workbook is a valid exported .xlsx file.", "danger")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        pw = request.form.get("password")
        if pw == os.environ.get("ADMIN_PASSWORD","admin123"):
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        flash("Invalid password","danger")
    return render_template("admin_login.html")

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("pledge_form"))

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    q = request.args.get("q", "").lower()
    all_pledges = Pledge.query.all()
    selected_year = request.args.get("year", "")
    pledges = all_pledges

    if q:
        pledges = [p for p in pledges if q in str(p.__dict__).lower()]

    if selected_year.isdigit():
        pledges = [p for p in pledges if p.pledge_date and p.pledge_date.year == int(selected_year)]

    total_bills = len(pledges)
    total_amount = round(sum(p.total_amount or 0 for p in pledges), 2)
    total_repaid_amount = round(sum(p.repayment_total_amount or 0 for p in pledges), 2)
    total_balance_amount = round(total_amount - total_repaid_amount, 2)
    total_grams = round(sum(p.total_grams or 0 for p in pledges), 2)
    total_returned_grams = round(sum(p.return_jewellery or 0 for p in pledges), 2)
    outstanding_grams = round(sum(p.balance_jewellery or 0 for p in pledges), 2)
    average_amount = round(total_amount / total_bills, 2) if total_bills else 0
    average_age = round(sum(p.age or 0 for p in pledges) / total_bills, 1) if total_bills else 0
    repayment_counts = {"Full": 0, "Partial": 0, "Nil": 0}
    for p in pledges:
        if p.repayment in repayment_counts:
            repayment_counts[p.repayment] += 1

    gender_counts = {"Male": 0, "Female": 0, "Other": 0}
    for p in pledges:
        if p.gender in gender_counts:
            gender_counts[p.gender] += 1

    monthly_counts = {}
    for pledge in pledges:
        if pledge.pledge_date:
            month_key = pledge.pledge_date.strftime("%Y-%m")
            monthly_counts[month_key] = monthly_counts.get(month_key, 0) + 1
    monthly_activity = [
        {"label": datetime.strptime(month, "%Y-%m").strftime("%b %Y"), "count": monthly_counts[month]}
        for month in sorted(monthly_counts)[-6:]
    ]
    max_monthly_count = max((item["count"] for item in monthly_activity), default=1)
    yearly_source = all_pledges
    if q:
        yearly_source = [p for p in yearly_source if q in str(p.__dict__).lower()]
    yearly_groups = {}
    for pledge in yearly_source:
        if pledge.pledge_date:
            year = pledge.pledge_date.year
            yearly_groups.setdefault(year, []).append(pledge)

    yearly_metrics = []
    for year in sorted(yearly_groups):
        year_records = yearly_groups[year]
        year_count = len(year_records)
        year_repayment = {"Full": 0, "Partial": 0, "Nil": 0}
        for pledge in year_records:
            if pledge.repayment in year_repayment:
                year_repayment[pledge.repayment] += 1
        yearly_metrics.append({
            "year": year,
            "count": year_count,
            "amount": round(sum(p.total_amount or 0 for p in year_records), 2),
            "repaid_amount": round(sum(p.repayment_total_amount or 0 for p in year_records), 2),
            "grams": round(sum(p.total_grams or 0 for p in year_records), 2),
            "returned": round(sum(p.return_jewellery or 0 for p in year_records), 2),
            "balance": round(sum(p.balance_jewellery or 0 for p in year_records), 2),
            "full": year_repayment["Full"],
            "partial": year_repayment["Partial"],
            "nil": year_repayment["Nil"]
        })

    year_options = sorted({item["year"] for item in yearly_metrics}, reverse=True)
    max_year_count = max((item["count"] for item in yearly_metrics), default=1)
    all_time_count = len(all_pledges)
    all_time_amount = round(sum(p.total_amount or 0 for p in all_pledges), 2)
    all_time_grams = round(sum(p.total_grams or 0 for p in all_pledges), 2)

    return render_template("admin_dashboard.html",
                           records=pledges,
                           total_bills=total_bills,
                           total_amount=total_amount,
                           total_repaid_amount=total_repaid_amount,
                           total_balance_amount=total_balance_amount,
                           total_grams=total_grams,
                           total_returned_grams=total_returned_grams,
                           outstanding_grams=outstanding_grams,
                           average_amount=average_amount,
                           average_age=average_age,
                           repayment_counts=repayment_counts,
                           gender_counts=gender_counts,
                           monthly_activity=monthly_activity,
                           max_monthly_count=max_monthly_count,
                           search_query=request.args.get("q", ""),
                           selected_year=selected_year,
                           year_options=year_options,
                           yearly_metrics=yearly_metrics,
                           max_year_count=max_year_count,
                           all_time_count=all_time_count,
                           all_time_amount=all_time_amount,
                           all_time_grams=all_time_grams)
@app.route("/admin/delete/<int:pledge_id>")
@admin_required
def admin_delete(pledge_id):
    pledge = Pledge.query.get_or_404(pledge_id)
    db.session.delete(pledge)
    db.session.commit()
    flash("Pledge deleted", "info")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/edit/<int:pledge_id>", methods=["GET", "POST"])
@admin_required
def admin_edit(pledge_id):
    pledge = Pledge.query.get_or_404(pledge_id)
    if request.method == "POST":
        pledge.gender = request.form.get("Gender")
        pledge.name = request.form.get("Name")
        pledge.age = request.form.get("Age") or None
        pledge.father_name = request.form.get("FatherName")
        pledge.family_name = request.form.get("FamilyName")
        pledge.kids_names = ",".join([
            request.form.get("KidsNames1", ""),
            request.form.get("KidsNames2", ""),
            request.form.get("KidsNames3", "")
        ])
        pledge.education = request.form.get("Education")
        pledge.occupation = request.form.get("Occupation")
        pledge.address = request.form.get("Address")
        pledge.phone = request.form.get("PhoneNumber")
        pledge.alt_number = request.form.get("AltNumber")
        pledge.aadhar = request.form.get("AadharNumber")
        pledge.hc_claim_form = request.form.get("HCClaimFormNumber")
        pledge.intro = request.form.get("Intro")
        pledge.num_ornaments = request.form.get("NumOrnaments") or None
        pledge.ornaments_details = str({k: v for k, v in request.form.items() if "Ornament" in k or "Grams" in k})
        pledge.pledge_date = datetime.strptime(request.form.get("PledgeDate"), "%Y-%m-%d").date() if request.form.get("PledgeDate") else None
        pledge.total_amount = float(request.form.get("TotalAmounts") or 0)
        pledge.total_grams = float(request.form.get("TotalGrams") or 0)
        pledge.return_jewellery = float(request.form.get("ReturnJewellery") or 0)
        pledge.balance_jewellery = float(request.form.get("BalanceJewellery") or 0)
        pledge.repayment = request.form.get("Repayment")
        pledge.repayment_details = str({k: v for k, v in request.form.items() if "Repay" in k})
        pledge.repayment_total_amount = round(sum(
            float(value or 0)
            for key, value in request.form.items()
            if key.startswith("RepayAmount")
        ), 2)
        pledge.remarks = request.form.get("Remarks")
        db.session.commit()
        flash("Pledge updated", "success")
        return redirect(url_for("admin_dashboard"))
    return render_template("form.html", data=pledge, edit=True)



if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1")
